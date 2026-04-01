import openpyxl
import cloudscraper
import requests
from bs4 import BeautifulSoup
import pypdf
import io
import os
import re
from urllib.parse import urlparse

EXCEL_FILE = r"C:\AI Projects\01_Apps\Somni\corpus\sources\CorpusDatabase1.0.xlsx"
SOURCES_DIR = r"C:\AI Projects\01_Apps\Somni\corpus\sources"
METADATA_DIR = r"C:\AI Projects\01_Apps\Somni\corpus\metadata"

scraper = cloudscraper.create_scraper(browser={
    'browser': 'chrome',
    'platform': 'windows',
    'desktop': True
})

def clean_filename(name):
    # Remove invalid characters
    return re.sub(r'[\\/*?:"<>|]', "", name).replace(" ", "_")

def process_html(url):
    response = scraper.get(url, timeout=15)
    response.raise_for_status()
    soup = BeautifulSoup(response.content, 'html.parser')
    
    # Remove scripts, styles, navigation, footers if possible
    for elem in soup(["script", "style", "nav", "footer", "header"]):
        elem.extract()
        
    text = soup.get_text(separator='\n\n')
    # Collapse multiple newlines
    text = re.sub(r'\n{3,}', '\n\n', text)
    title = soup.title.string.strip() if soup.title else url.split('/')[-1]
    return title, text.strip()

def process_pdf(url):
    response = scraper.get(url, timeout=20)
    response.raise_for_status()
    pdf_file = io.BytesIO(response.content)
    reader = pypdf.PdfReader(pdf_file)
    text = []
    for page in reader.pages:
        text.append(page.extract_text())
    title = url.split('/')[-1]
    title = title.replace('.pdf', '')
    return title, '\n\n'.join(text).strip()

def main():
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    
    source_index = ["# Corpus Source Index\n\n| Organisation | Title | Status | URL |\n|---|---|---|---|"]
    
    success_count = 0
    fail_count = 0
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        org = row[0].value
        url = row[1].value
        
        if not org or not url or not str(url).startswith('http'):
            continue
            
        # Skip homepages (path elements <= 2)
        parsed = urlparse(url)
        path = parsed.path.rstrip('/')
        if not path or path.count('/') <= 1:
            continue
            
        print(f"Processing: {url}")
        
        try:
            if '.pdf' in url.lower():
                title, content = process_pdf(url)
            else:
                title, content = process_html(url)
                
            safe_org = clean_filename(org)
            safe_title = clean_filename(title[:50]) # limit length
            filename = f"{safe_org}__{safe_title}.md"
            filepath = os.path.join(SOURCES_DIR, filename)
            
            md_content = f"# Source: {title}\n\n- **URL**: {url}\n- **Organisation**: {org}\n- **Date Accessed**: 2026-04-01\n\n## Content\n\n{content}"
            
            with open(filepath, 'w', encoding='utf-8') as f:
                f.write(md_content)
                
            source_index.append(f"| {org} | {title} | 🟢 Processed | {url} |")
            success_count += 1
            
        except requests.exceptions.HTTPError as e:
            if e.response.status_code == 403:
                print(f"403 Forbidden for {url} - Needs manual extraction")
                source_index.append(f"| {org} | Unknown | 🔴 403 Forbidden | {url} |")
            elif e.response.status_code == 404:
                print(f"404 Not Found for {url}")
                source_index.append(f"| {org} | Unknown | 🔴 404 Not Found | {url} |")
            else:
                print(f"HTTP error for {url}: {e}")
                source_index.append(f"| {org} | Unknown | 🔴 Error HTTP {e.response.status_code} | {url} |")
            fail_count += 1
        except Exception as e:
            print(f"Failed to process {url}: {str(e)}")
            source_index.append(f"| {org} | Unknown | 🔴 Error | {url} |")
            fail_count += 1
            
    # Write index
    index_path = os.path.join(METADATA_DIR, "source_index.md")
    with open(index_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(source_index))
        
    print(f"\nDone! Successfully extracted {success_count} sources. {fail_count} failed.")

if __name__ == "__main__":
    main()
