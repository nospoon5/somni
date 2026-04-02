import openpyxl

EXCEL_FILE = r"C:\AI Projects\01_Apps\Somni\corpus\sources\CorpusDatabase1.0.xlsx"

new_urls = [
    ("Sleep Foundation", "https://www.sleepfoundation.org/baby-sleep/ferber-method"),
    ("Huckleberry", "https://huckleberrycare.com/blog/ferber-method-for-sleep-training-what-age-to-start"),
    ("BabyCenter", "https://www.babycenter.com/baby/sleep/how-to-try-the-ferber-method-of-sleep-training-for-your-baby_7755"),
    ("Happiest Baby", "https://www.happiestbaby.com/blogs/baby/sleep-training?srsltid=AfmBOorAhI0CpA0edV3-AdBJ6FHmRLWx94IEb4c4dZeEkMWrYBViK8Nu"),
    ("Huckleberry", "https://huckleberrycare.com/blog/cry-it-out-method-aka-extinction-method-is-it-right-for-your-baby"),
    ("Pampers", "https://www.pampers.com/en-us/baby/sleep/article/cry-it-out-method"),
    ("Dr Craig Canapari", "https://drcraigcanapari.com/cry-it-out-sleep-training-explained-how-to-use-cio-to-sleep-train-a-baby/"),
    ("Happiest Baby", "https://www.happiestbaby.com/blogs/baby/avoid-crying-it-out-sleep-training?srsltid=AfmBOoqgzQx-N4ANfwOZJ08Yqe5AR2s33cZDa6LEgsOPpngVH1NiDGYT")
]

wb = openpyxl.load_workbook(EXCEL_FILE)
ws = wb.active

# Check if already added
existing_urls = set()
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    if row[1].value:
        existing_urls.add(row[1].value)

added = 0
for org, url in new_urls:
    if url not in existing_urls:
        ws.append([org, url])
        added += 1

wb.save(EXCEL_FILE)
print(f"Added {added} new URLs to Excel file.")
